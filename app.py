from flask import Flask, render_template, request, redirect, url_for, jsonify, abort, send_file

import json
import os
import urllib.parse
from datetime import date, timedelta, datetime
import subprocess
import re

from flask import jsonify

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


app = Flask(__name__)

FILE_NAME = "catalog.json"
orders = {}  # sku -> qty
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN")

TYPE_ORDER = [
    "Cups",
    "Lids",
    "Paper Goods",
    "Containers",
    "Cleaning Products",
    "Other"
]

CUP_SUBTYPE_ORDER = {
    "paper": 0,
    "plastic": 1,
    "portion": 2
}

WASTE_FILE = "waste_logs.json"
PASTRY_PRICES_FILE = "pastry_prices.json"

WASTE_REASONS = [
    "Not sold",
    "Overproduced",
    "Expired",
    "Damaged",
    "Staff error",
    "Other",
]





# ---------- Helpers ----------

def load_catalog():
    if not os.path.exists(FILE_NAME):
        return []
    with open(FILE_NAME, "r", encoding="utf-8") as f:
        return json.load(f)

def save_catalog(catalog):
    with open("catalog.json", "w") as f:
        json.dump(catalog, f, indent=2)

    subprocess.run(["git", "add", "catalog.json"])
    subprocess.run(["git", "commit", "-m", "Update catalog"])
    subprocess.run([
        "git",
        "push",
        f"https://{GITHUB_TOKEN}@github.com/joacotol/redchurch_inventory_system.git",
        "main"
    ])

def format_day_with_suffix(d):
    if 11 <= d.day <= 13:
        suffix = "th"
    else:
        suffix = {1: "st", 2: "nd", 3: "rd"}.get(d.day % 10, "th")
    return f"{d.strftime('%B')} {d.day}{suffix}"



# - For sorting the cups
def cup_subtype(name: str) -> int:
    n = name.lower()

    if "portion" in n:
        return 2
    if "clear" in n or "plastic" in n:
        return 1
    return 0  # paper (default)

def extract_oz(name: str) -> int:
    match = re.search(r"(\d+)\s?oz", name.lower())
    return int(match.group(1)) if match else 999


# -- WASTE LOG HELPERS
def parse_iso_date(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None

def monday_of_week(d: date) -> date:
    return d - timedelta(days=d.weekday())

def display_full_date(d: date) -> str:
    # You already have format_day_with_suffix() in your file
    try:
        return f"{format_day_with_suffix(d)}, {d.year}"
    except Exception:
        return d.isoformat()

def git_push_file_if_possible(file_path: str, message: str):
    # Keeps your app from crashing if push fails in production
    if not GITHUB_TOKEN:
        return
    try:
        subprocess.run(["git", "add", file_path], check=False)
        subprocess.run(["git", "commit", "-m", message], check=False)
        subprocess.run(
            [
                "git",
                "push",
                f"https://{GITHUB_TOKEN}@github.com/joacotol/redchurch_inventory_system.git",
                "main",
            ],
            check=False,
        )
    except Exception as e:
        print(f"[WARN] git push failed: {e}")

def load_pastry_prices():
    if not os.path.exists(PASTRY_PRICES_FILE):
        # If file doesn't exist, start empty (or create default if you prefer)
        return []

    with open(PASTRY_PRICES_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Normalize
    cleaned = []
    for row in data if isinstance(data, list) else []:
        name = str(row.get("name", "")).strip()
        price = row.get("price", None)
        try:
            price = float(price) if price is not None else None
        except Exception:
            price = None
        if name:
            cleaned.append({"name": name, "price": price if price is not None else 0.0})
    return cleaned

def save_pastry_prices(items: list):
    with open(PASTRY_PRICES_FILE, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)
    git_push_file_if_possible(PASTRY_PRICES_FILE, "Update pastry prices")

def pastry_items_and_price_map():
    items = load_pastry_prices()
    names = [x["name"] for x in items]
    price_map = {x["name"]: float(x["price"]) for x in items}
    return names, price_map

def load_waste_logs():
    if not os.path.exists(WASTE_FILE):
        return {}
    with open(WASTE_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_waste_logs(logs: dict, message: str):
    with open(WASTE_FILE, "w", encoding="utf-8") as f:
        json.dump(logs, f, indent=2, ensure_ascii=False)
    git_push_file_if_possible(WASTE_FILE, message)

def build_date_options(logs: dict, include_today: bool = True, limit: int = 60):
    keys = []
    for k in logs.keys():
        if parse_iso_date(k):
            keys.append(k)
    keys = sorted(keys, reverse=True)

    if include_today:
        t = date.today().isoformat()
        if t not in keys:
            keys = [t] + keys

    keys = keys[:limit]
    opts = []
    for iso in keys:
        d = parse_iso_date(iso)
        opts.append({"iso": iso, "label": display_full_date(d) if d else iso})
    return opts

# -- WEEKLY SUMMARY HELPERS
def aggregate_week(logs: dict, start_date: date, current_price_map: dict):
    total_qty = 0
    total_cost = 0.0
    item_map = {}   # item -> {qty, cost}
    daily = []      # [{label, iso, qty, cost}]
    unknown_price_items = set()

    for i in range(7):
        d = start_date + timedelta(days=i)
        iso = d.isoformat()

        day_entries = []
        if isinstance(logs.get(iso, {}), dict):
            day_entries = (logs.get(iso, {}) or {}).get("entries", []) or []

        day_qty = 0
        day_cost = 0.0

        for e in day_entries:
            item = str(e.get("item", "")).strip()
            try:
                qty = int(e.get("qty", 0))
            except Exception:
                qty = 0

            unit_price = e.get("unit_price", None)
            if unit_price is None:
                unit_price = current_price_map.get(item)

            try:
                unit_price = float(unit_price) if unit_price is not None else None
            except Exception:
                unit_price = None

            if unit_price is None:
                unknown_price_items.add(item)

            cost = (qty * unit_price) if unit_price is not None else 0.0

            total_qty += qty
            total_cost += cost
            day_qty += qty
            day_cost += cost

            it = item_map.setdefault(item, {"qty": 0, "cost": 0.0})
            it["qty"] += qty
            it["cost"] += cost

        daily.append(
            {
                "iso": iso,
                "label": display_full_date(d),
                "qty": day_qty,
                "cost": round(day_cost, 2),
            }
        )

    items = [
        {"item": k, "qty": v["qty"], "cost": round(v["cost"], 2)}
        for k, v in item_map.items()
        if k
    ]
    items.sort(key=lambda x: (x["cost"], x["qty"]), reverse=True)

    return {
        "total_qty": total_qty,
        "total_cost": round(total_cost, 2),
        "items": items,
        "item_map": item_map,
        "daily": daily,
        "unknown_price_items": sorted([x for x in unknown_price_items if x]),
    }

def weekly_waste_aggregate_for_export(start_date: date):
    """
    Collect ONLY days in the week that actually logged waste entries (and qty > 0).
    Uses stored unit_price if present; falls back to current pastry_prices.json mapping.
    """
    logs = load_waste_logs()
    _, price_map = pastry_items_and_price_map()

    daily_rows = []        # days with logged waste
    item_totals = {}       # item -> {qty, cost}
    raw_entries = []       # flat list for Entries sheet
    missing_price_items = set()

    for i in range(7):
        d = start_date + timedelta(days=i)
        iso = d.isoformat()
        weekday = d.strftime("%A")

        day_obj = logs.get(iso, {})
        day_entries = (day_obj.get("entries", []) if isinstance(day_obj, dict) else []) or []

        day_qty = 0
        day_cost = 0.0

        for e in day_entries:
            item = str(e.get("item", "")).strip()
            reason = str(e.get("reason", "")).strip() or "Other"

            try:
                qty = int(e.get("qty", 0))
            except Exception:
                qty = 0

            if not item or qty <= 0:
                continue

            unit_price = e.get("unit_price", None)
            if unit_price is None:
                unit_price = price_map.get(item)

            try:
                unit_price = float(unit_price) if unit_price is not None else None
            except Exception:
                unit_price = None

            if unit_price is None:
                missing_price_items.add(item)

            cost = (qty * unit_price) if unit_price is not None else 0.0

            day_qty += qty
            day_cost += cost

            it = item_totals.setdefault(item, {"qty": 0, "cost": 0.0})
            it["qty"] += qty
            it["cost"] += cost

            raw_entries.append({
                "date": iso,
                "day": weekday,
                "item": item,
                "reason": reason,
                "qty": qty,
                "unit_price": unit_price,
                "cost": round(cost, 2),
            })

        # IMPORTANT: only include days that actually logged any waste
        if day_qty > 0:
            daily_rows.append({
                "date": iso,
                "day": weekday,
                "qty": day_qty,
                "cost": round(day_cost, 2),
            })

    items_sorted = [{"item": k, "qty": v["qty"], "cost": round(v["cost"], 2)} for k, v in item_totals.items()]
    items_sorted.sort(key=lambda x: (x["cost"], x["qty"]), reverse=True)

    total_qty = sum(r["qty"] for r in daily_rows)
    total_cost = round(sum(r["cost"] for r in daily_rows), 2)

    return {
        "daily": daily_rows,
        "items": items_sorted,
        "entries": raw_entries,
        "total_qty": total_qty,
        "total_cost": total_cost,
        "missing_price_items": sorted([x for x in missing_price_items if x]),
    }


def build_weekly_waste_workbook(start_date: date, agg: dict):
    end_date = start_date + timedelta(days=6)
    week_label = f"{start_date.isoformat()} to {end_date.isoformat()}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    # --- styles
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")  # dark slate
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # --- header
    ws["A1"] = "Weekly Pastry Waste Report"
    ws["A1"].font = title_font
    ws["A2"] = f"Week: {week_label}"
    ws["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # --- totals
    ws["A5"] = "Totals"
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "Total units wasted"
    ws["B6"] = agg["total_qty"]

    ws["A7"] = "Total waste cost"
    ws["B7"] = agg["total_cost"]
    ws["B7"].number_format = '"$"#,##0.00'

    row = 9
    if agg["missing_price_items"]:
        ws["A9"] = "Missing prices (cost shown as $0.00 for these items):"
        ws["A9"].font = Font(bold=True, color="B91C1C")
        ws["A10"] = ", ".join(agg["missing_price_items"])
        row = 12

    # --- Daily table
    ws[f"A{row}"] = "Cost by Day (only days with logged waste)"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    daily_header = ["Date", "Day", "Units", "Cost"]
    for c, h in enumerate(daily_header, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row += 1
    daily_start_row = row
    for r in agg["daily"]:
        ws.cell(row=row, column=1, value=r["date"])
        ws.cell(row=row, column=2, value=r["day"])
        ws.cell(row=row, column=3, value=r["qty"]).alignment = right
        ccell = ws.cell(row=row, column=4, value=r["cost"])
        ccell.number_format = '"$"#,##0.00'
        ccell.alignment = right
        row += 1
    daily_end_row = row - 1

    row += 2

    # --- Top items table
    ws[f"A{row}"] = "Top Items by Cost"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    item_header = ["Item", "Units", "Cost"]
    for c, h in enumerate(item_header, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    row += 1
    item_start_row = row
    top_items = agg["items"][:15]
    for it in top_items:
        ws.cell(row=row, column=1, value=it["item"])
        ws.cell(row=row, column=2, value=it["qty"]).alignment = right
        ccell = ws.cell(row=row, column=3, value=it["cost"])
        ccell.number_format = '"$"#,##0.00'
        ccell.alignment = right
        row += 1
    item_end_row = row - 1

    # widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 14

    # charts (Excel-native; Google Sheets may or may not convert them automatically)
    if daily_end_row >= daily_start_row:
        chart1 = BarChart()
        chart1.title = "Cost by Day"
        chart1.y_axis.title = "Cost ($)"

        data = Reference(ws, min_col=4, min_row=daily_start_row-1, max_row=daily_end_row)
        cats = Reference(ws, min_col=2, min_row=daily_start_row, max_row=daily_end_row)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 8
        chart1.width = 18
        ws.add_chart(chart1, "F12")

    if item_end_row >= item_start_row:
        chart2 = BarChart()
        chart2.title = "Top Items by Cost"
        chart2.y_axis.title = "Cost ($)"

        data2 = Reference(ws, min_col=3, min_row=item_start_row-1, max_row=item_end_row)
        cats2 = Reference(ws, min_col=1, min_row=item_start_row, max_row=item_end_row)
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(cats2)
        chart2.height = 10
        chart2.width = 18
        ws.add_chart(chart2, "F32")

    # --- Entries sheet (filterable)
    ws2 = wb.create_sheet("Entries")
    headers = ["Date", "Day", "Item", "Reason", "Qty", "Unit Price", "Cost"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    r = 2
    for e in agg["entries"]:
        ws2.cell(row=r, column=1, value=e["date"])
        ws2.cell(row=r, column=2, value=e["day"])
        ws2.cell(row=r, column=3, value=e["item"])
        ws2.cell(row=r, column=4, value=e["reason"])
        ws2.cell(row=r, column=5, value=e["qty"]).alignment = right

        up = ws2.cell(row=r, column=6, value=(e["unit_price"] if e["unit_price"] is not None else ""))
        if e["unit_price"] is not None:
            up.number_format = '"$"#,##0.00'
        up.alignment = right

        cc = ws2.cell(row=r, column=7, value=e["cost"])
        cc.number_format = '"$"#,##0.00'
        cc.alignment = right
        r += 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 28
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 8
    ws2.column_dimensions["F"].width = 12
    ws2.column_dimensions["G"].width = 12
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:G{max(1, r-1)}"

    return wb

# ---------- Routes ----------

@app.route("/", methods=["GET"])
def index():
    catalog = load_catalog()

    def sort_key(item):
        try:
            type_index = TYPE_ORDER.index(item.get("type", "Other"))
        except ValueError:
            type_index = len(TYPE_ORDER)

        name_for_sort = (item.get("display_name") or item["name"]).lower()

        # Special sorting for Cups
        if item.get("type") == "Cups":
            return (
                type_index,
                cup_subtype(name_for_sort),
                extract_oz(name_for_sort),
                name_for_sort
            )

        # Default sorting for all other types
        return (type_index, name_for_sort)


    catalog = sorted(catalog, key=sort_key)

    return render_template(
        "index.html",
        items=catalog,
        orders=orders,
        product_types=TYPE_ORDER
    )



@app.route("/add_item", methods=["POST"])
def add_item():
    catalog = load_catalog()
    item_type = request.form["type"]

    if item_type not in TYPE_ORDER:
        abort(400, "Invalid Product Type")

    catalog.append({
        "sku": request.form["sku"],
        "name": request.form["name"],
        "unit": request.form["unit"],
        "type": item_type
    })

    save_catalog(catalog)
    return redirect(url_for("index"))


@app.route("/add_to_order", methods=["POST"])
def add_to_order():
    sku = request.form["sku"]
    qty = int(request.form["qty"])
    orders[sku] = orders.get(sku, 0) + qty

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(success=True)

    return redirect(url_for("index"))

@app.route("/remove_from_order", methods=["POST"])
def remove_from_order():
    orders.pop(request.form["sku"], None)

    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify(success=True)

    return redirect(url_for("index"))

@app.route("/email")
def email_order():
    catalog = load_catalog()
    today = date.today().strftime("%B %d")

    subject = f"Redchurch Cafe Weekly Order – {today}"

    today_date = date.today()
    delivery_date = today_date + timedelta(days=1)

    today_formatted = format_day_with_suffix(today_date)
    delivery_formatted = format_day_with_suffix(delivery_date)
    delivery_weekday = delivery_date.strftime("%A")


    lines = []
    for item in catalog:
        if item["sku"] in orders:
            lines.append(
                f"{orders[item['sku']]} {item['unit']}(s) – "
                f"[{item['sku']}] – {item['name']}"
            )

    newline = "\r\n"

    body = (
        f"Good morning!{newline}"
        f"This is our order for the week of {today_formatted}, "
        f"for a delivery of {delivery_weekday} {delivery_formatted}, please."
        f"{newline}{newline}"
        + newline.join(lines)
        + f"{newline}{newline}"
        f"Thank you,{newline}"
        f"Stefanie Forget{newline}"
        f"Manager{newline}"
        f"Redchurch Cafe{newline}"
        f"68 King Street E, Hamilton ON"
    )



    gmail_url = (
        "https://mail.google.com/mail/?view=cm&fs=1&tf=1"
        f"&su={urllib.parse.quote(subject)}"
        f"&body={urllib.parse.quote(body)}"
    )

    mailto_url = (
        "mailto:?"
        f"subject={urllib.parse.quote(subject)}"
        f"&body={urllib.parse.quote(body)}"
    )

    return jsonify({
        "gmail": gmail_url,
        "mailto": mailto_url
    })

@app.route("/order_summary")
def order_summary():
    catalog = load_catalog()
    summary = []

    for item in catalog:
        if item["sku"] in orders:
            summary.append({
                "sku": item["sku"],
                "name": item["name"],
                "unit": item["unit"],
                "qty": orders[item["sku"]]
            })

    return jsonify(summary)

# -- WASTE LOG ROUTES
@app.route("/waste", methods=["GET"])
def waste_log():
    logs = load_waste_logs()
    pastry_items, price_map = pastry_items_and_price_map()

    requested = (request.args.get("date") or "").strip()
    selected_date = parse_iso_date(requested) if requested else date.today()
    if not selected_date:
        selected_date = date.today()

    selected_iso = selected_date.isoformat()
    today_display = display_full_date(selected_date)

    existing = logs.get(selected_iso, {})
    entries = existing.get("entries", []) if isinstance(existing, dict) else []
    if not entries:
        entries = [{"item": "", "qty": 1, "reason": WASTE_REASONS[0]}]

    date_options = build_date_options(logs, include_today=True)
    week_start = monday_of_week(selected_date)

    return render_template(
        "waste.html",
        today=today_display,
        today_iso=selected_iso,
        pastry_items=pastry_items,
        waste_reasons=WASTE_REASONS,
        initial_rows=entries,
        date_options=date_options,
        selected_date_iso=selected_iso,
        week_start_iso=week_start.isoformat(),
    )


@app.route("/waste/save", methods=["POST"])
def waste_save():
    payload = request.get_json(silent=True) or {}
    date_iso = str(payload.get("date", "")).strip()
    entries = payload.get("entries", [])

    d = parse_iso_date(date_iso)
    if not d:
        abort(400, "Invalid date. Expected YYYY-MM-DD")

    if not isinstance(entries, list):
        abort(400, "Invalid entries")

    _, price_map = pastry_items_and_price_map()

    cleaned = []
    for e in entries:
        if not isinstance(e, dict):
            continue

        item = str(e.get("item", "")).strip()
        reason = str(e.get("reason", "")).strip() or "Other"

        try:
            qty = int(e.get("qty", 0))
        except Exception:
            qty = 0

        if not item or qty <= 0:
            continue

        if reason not in WASTE_REASONS:
            reason = "Other"

        unit_price = price_map.get(item)  # may be missing if no price configured

        cleaned.append(
            {"item": item, "qty": qty, "reason": reason, "unit_price": unit_price}
        )

    logs = load_waste_logs()
    logs[date_iso] = {
        "date": date_iso,
        "entries": cleaned,
        "updated_at": datetime.utcnow().isoformat() + "Z",
    }

    save_waste_logs(logs, f"Waste log {date_iso}")
    return jsonify(success=True, saved=len(cleaned))


@app.route("/waste/weekly", methods=["GET"])
def waste_weekly():
    logs = load_waste_logs()
    _, price_map = pastry_items_and_price_map()

    start_str = (request.args.get("start") or "").strip()
    start_date = parse_iso_date(start_str) if start_str else monday_of_week(date.today())
    if not start_date:
        start_date = monday_of_week(date.today())

    start_date = monday_of_week(start_date)
    end_date = start_date + timedelta(days=6)

    # Build week dropdown (last 12 weeks)
    this_monday = monday_of_week(date.today())
    week_options = []
    for w in range(12):
        sd = this_monday - timedelta(days=7 * w)
        week_options.append({"iso": sd.isoformat(), "label": f"Week of {display_full_date(sd)}"})

    curr = aggregate_week(logs, start_date, price_map)
    prev_start = start_date - timedelta(days=7)
    prev = aggregate_week(logs, prev_start, price_map)

    # Trend vs last week
    delta_qty = curr["total_qty"] - prev["total_qty"]
    delta_cost = round(curr["total_cost"] - prev["total_cost"], 2)

    def pct_change(curr_val, prev_val):
        if prev_val == 0:
            return None
        return round(((curr_val - prev_val) / prev_val) * 100, 1)

    pct_qty = pct_change(curr["total_qty"], prev["total_qty"])
    pct_cost = pct_change(curr["total_cost"], prev["total_cost"])

    # Top 3 items by cost + last week comparison
    top3 = []
    for row in curr["items"][:3]:
        item = row["item"]
        prev_cost = round(prev["item_map"].get(item, {}).get("cost", 0.0), 2)
        dcost = round(row["cost"] - prev_cost, 2)
        pc = pct_change(row["cost"], prev_cost)
        top3.append({
            "item": item,
            "qty": row["qty"],
            "cost": row["cost"],
            "prev_cost": prev_cost,
            "delta_cost": dcost,
            "pct_cost": pc,
        })

    # Chart data
    chart_daily_labels = [d["label"] for d in curr["daily"]]
    chart_daily_costs = [d["cost"] for d in curr["daily"]]

    top_items_for_chart = curr["items"][:5]
    chart_item_labels = [x["item"] for x in top_items_for_chart]
    chart_item_costs = [x["cost"] for x in top_items_for_chart]

    return render_template(
        "waste_weekly.html",
        start_iso=start_date.isoformat(),
        end_iso=end_date.isoformat(),
        start_label=display_full_date(start_date),
        end_label=display_full_date(end_date),
        week_options=week_options,
        selected_start_iso=start_date.isoformat(),

        total_qty=curr["total_qty"],
        total_cost=curr["total_cost"],

        prev_total_qty=prev["total_qty"],
        prev_total_cost=prev["total_cost"],
        delta_qty=delta_qty,
        delta_cost=delta_cost,
        pct_qty=pct_qty,
        pct_cost=pct_cost,

        top3=top3,
        daily=curr["daily"],
        items=curr["items"],
        unknown_price_items=curr["unknown_price_items"],

        chart_daily_labels=chart_daily_labels,
        chart_daily_costs=chart_daily_costs,
        chart_item_labels=chart_item_labels,
        chart_item_costs=chart_item_costs,
    )

@app.route("/waste/weekly/export", methods=["GET"])
def export_waste_weekly():
    start_str = (request.args.get("start") or "").strip()
    start_date = parse_iso_date(start_str) if start_str else monday_of_week(date.today())
    if not start_date:
        start_date = monday_of_week(date.today())

    start_date = monday_of_week(start_date)

    agg = weekly_waste_aggregate_for_export(start_date)
    wb = build_weekly_waste_workbook(start_date, agg)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    end_date = start_date + timedelta(days=6)
    filename = f"weekly_waste_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@app.route("/waste/prices", methods=["GET"])
def waste_prices():
    items = load_pastry_prices()
    return render_template("waste_prices.html", items=items)

@app.route("/waste/prices/save", methods=["POST"])
def waste_prices_save():
    payload = request.get_json(silent=True) or {}
    items = payload.get("items", [])

    if not isinstance(items, list):
        abort(400, "Invalid payload")

    cleaned = []
    for r in items:
        if not isinstance(r, dict):
            continue
        name = str(r.get("name", "")).strip()
        if not name:
            continue
        try:
            price = float(r.get("price", 0))
        except Exception:
            price = 0.0
        if price < 0:
            price = 0.0
        cleaned.append({"name": name, "price": round(price, 2)})

    # Optional: sort A→Z
    cleaned.sort(key=lambda x: x["name"].lower())

    save_pastry_prices(cleaned)
    return jsonify(success=True, count=len(cleaned))



if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000, debug=False)
